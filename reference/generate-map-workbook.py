# -*- coding: utf-8 -*-
"""
Generates reference/module-feature-permission-map.xlsx — the dev lookup workbook:
  1. Read me
  2. Module - Feature - Permission   (product module -> feature -> v2 permission tags,
                                      with role/grant defaults and legacy DB sources)
  3. Role x Permission               (the 5 fixed roles)
  4. Grant x Permission              (the 4 admin grants)
  5. Legacy tag audit                (all 121 DB permission tags; flags every tag
                                      that does NOT carry into v2)

Sources (never hand-edit the workbook — rerun this script):
  - index.html            SETS / ROLES / GRANTS / MODULES / PERMMOD (extracted live via node)
  - coverage-map.csv      the authoritative 121 legacy-tag -> v2-home table
  - data/Permissions.json + Features.json + Modules.json  (optional, local-only DB export:
                          adds real DB names to the audit sheet; falls back to the CSV alone)

Run from the repo root:  python reference/generate-map-workbook.py
"""
import csv, json, os, subprocess, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "reference", "module-feature-permission-map.xlsx")

# ---------------------------------------------------------------- model from index.html
NODE_SNIPPET = r"""
function mkEl(){return{style:{},innerHTML:'',textContent:'',disabled:false,className:'',value:'',options:{length:0},
 classList:{toggle(){},add(){},remove(){},contains:()=>false},appendChild(){},remove(){}};}
const els={};
global.document={getElementById:id=>els[id]||(els[id]=mkEl()),querySelectorAll:()=>[mkEl(),mkEl(),mkEl(),mkEl()],
 querySelector:()=>mkEl(),createElement:()=>mkEl(),body:{appendChild(){}}};
global.window={scrollTo(){}};
const fs=require('fs');
const script=fs.readFileSync(process.argv[1]+'/index.html','utf8').match(/<script>([\s\S]*)<\/script>/)[1];
const fn=new Function(script+';return {SETS,ROLES,GRANTS,GRANTNAMES,MODULES,PERMMOD};');
process.stdout.write(JSON.stringify(fn()));
"""

def load_model():
    out = subprocess.run(["node", "-e", NODE_SNIPPET, ROOT],
                         capture_output=True, text=True, encoding="utf-8")
    if out.returncode != 0:
        sys.exit("node extraction failed:\n" + out.stderr)
    return json.loads(out.stdout)

M = load_model()
SETS, ROLES, GRANTS = M["SETS"], M["ROLES"], M["GRANTS"]
GRANTNAMES, MODULES, PERMMOD = M["GRANTNAMES"], M["MODULES"], M["PERMMOD"]
set_by_id = {s["id"]: s for s in SETS}
mod_by_id = {m["id"]: m for m in MODULES}
ROLE_ORDER = ["l1", "l3", "l4", "regular", "senior"]
GRANT_ORDER = ["people", "tech", "fullsite", "global"]

def perm_name(tag):
    sid, pid = tag.split(".")
    return next(p["n"] for p in set_by_id[sid]["perms"] if p["id"] == pid)

def in_role(role_id, tag):
    sid = tag.split(".")[0]
    return sid != "flags" and sid in ROLES[role_id]["std"]   # flags are never a role default

def in_grant(grant_id, tag):
    return tag.split(".")[0] in GRANTS[grant_id]             # global carries all 10 sets incl. flags

ALL_TAGS = [s["id"] + "." + p["id"] for s in SETS for p in s["perms"]]

# ------------------------------------------------- curated feature layer, per product module
FEATURES = {
 "core": [
   ("Plant dashboard & issue log", ["readplant.dash"]),
   ("Users & access management", ["people.invite", "people.grants", "people.groups",
                                  "people.skills", "people.notif", "people.audit"]),
   ("Plant & site setup", ["tech.plant", "tech.sitetpl"]),
   ("Global governance (DigitalPaani staff)", ["templates.library", "templates.publish", "templates.onboard"]),
   ("View-as / impersonation — EXCEPTION grant", ["flags.impersonate"]),
 ],
 "ops": [
   ("Issue sessions & guided diagnostics", ["work.sessions", "work.rc", "work.media",
                                            "work.raise", "work.handoff"]),
   ("Approvals & closure", ["approve.gates", "approve.selfapprove", "approve.forceclose",
                            "approve.reopen", "approve.override"]),
   ("Multi-plant ops oversight", ["oversight.multiws", "oversight.cosign", "oversight.deadline"]),
   ("Issue rules, flows & escalation config", ["tech.rules", "tech.flows", "tech.routing"]),
 ],
 "tasks": [
   ("Everyday tasks, shifts & history", ["readplant.tasks", "readplant.myshift", "readplant.history"]),
   ("Task & insight assignment", ["approve.assign", "approve.shiftview"]),
   ("Rosters & schedules", ["people.rosters"]),
 ],
 "data": [
   ("Data, lab, water-quality & logbook entry", ["readplant.data"]),
   ("Data correction (supervisor)", ["approve.datacorrect"]),
   ("Site analytics formulas", ["tech.formulas"]),
 ],
 "analytics": [
   ("Plant analytics views", ["readplant.layout", "readplant.insights", "readplant.reports"]),
   ("Portfolio views", ["portfolio.overview", "portfolio.compare"]),
   ("Dashboard configuration", ["tech.dashcfg"]),
 ],
 "iot": [
   ("Remote actuation — SENSITIVE, EXCEPTION grant", ["remote.actuate"]),
   ("Sensors & hardware configuration", ["tech.sensors", "tech.iot"]),
   ("Sensor health dashboard — EXCEPTION grant", ["flags.sensorhealth"]),
 ],
 "inv": [
   ("Stock operations (operator)", ["work.inventory"]),
   ("Movement log (supervisor)", ["approve.invlogs"]),
   ("Store & catalog setup", ["tech.stores"]),
 ],
}
covered = [t for feats in FEATURES.values() for _, tags in feats for t in tags]
assert sorted(covered) == sorted(ALL_TAGS), \
    "feature map out of sync with SETS: " + str(set(covered) ^ set(ALL_TAGS))
assert all(PERMMOD[t] == mod for mod, feats in FEATURES.items() for _, tags in feats for t in tags), \
    "a permission is filed under the wrong module"

# ---------------------------------------------------------------- coverage map + DB export
def read_coverage():
    rows = []
    with open(os.path.join(ROOT, "coverage-map.csv"), encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append({k.strip(): (v or "").strip() for k, v in r.items()})
    return rows

COVER = read_coverage()
legacy_sources = {}                                   # v2 tag -> [legacy tags]
for r in COVER:
    home = r["V2 home"]
    if home in PERMMOD:
        legacy_sources.setdefault(home, []).append(r["Production permission"])

def read_db():
    """Optional local DB export — real names for the audit sheet."""
    base = os.path.join(ROOT, "data")
    alt = "D:/Roles/data"                             # worktrees don't carry untracked data/
    for d in (base, alt):
        p = os.path.join(d, "Permissions.json")
        if os.path.exists(p):
            perms = json.load(open(p, encoding="utf-8"))
            feats = {f["featureTag"]: f["name"] for f in json.load(open(os.path.join(d, "Features.json"), encoding="utf-8"))}
            mods = {m["moduleTag"]: m["name"] for m in json.load(open(os.path.join(d, "Modules.json"), encoding="utf-8"))}
            return perms, feats, mods, d
    return None, {}, {}, None

DB_PERMS, DB_FEATS, DB_MODS, DB_DIR = read_db()

# --------------------------------------------------------------------------- build workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY, TEAL_TINT, AMBER_BG, AMBER_INK, GREY = "002454", "E8F5F6", "FFF8E1", "B45309", "8A95A3"
def F(**kw):
    kw.setdefault("name", "Arial")
    kw.setdefault("size", 10)
    return Font(**kw)
HDR_FILL = PatternFill("solid", fgColor=NAVY)
FEAT_FILL = PatternFill("solid", fgColor=TEAL_TINT)
FLAG_FILL = PatternFill("solid", fgColor=AMBER_BG)
THIN = Border(*[Side(style="thin", color="D5DBE2")] * 4)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F(bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------- 1 · Read me
ws = wb.active
ws.title = "Read me"
set_widths(ws, [118])
lines = [
 ("User Center — Roles & Permissions (v2): module → feature → permission → role/grant lookup", dict(bold=True, size=14, color=NAVY)),
 ("Generated 2026-08-13 · owner: Mihir Sethi (APM, DigitalPaani) · read-only reference — to change anything, change the model and rerun reference/generate-map-workbook.py", dict(size=9, color=GREY)),
 ("", {}),
 ("THE ONE RULE:  effective access at a plant = the person's permission AND the plant's licensed module. Modules cap; roles default; per-person reasoned exceptions flex.", dict(bold=True)),
 ("", {}),
 ("Decisions this workbook already reflects:", dict(bold=True, color=NAVY)),
 ("  • One role per user, account-wide (2026-07-30). A role describes the person, not a plant — there is no per-plant role anywhere.", {}),
 ("  • Custom roles are REMOVED (2026-08-13). None will be created; no admin — People Admins included — can create or apply one. The 5 roles + 4 grants in this workbook are the complete, permanent role vocabulary.", {}),
 ("  • No bulk role changes. Plant-wide bulk = add people to a plant's access list + add/remove individual permissions as per-person exceptions.", {}),
 ("", {}),
 ("Sheets:", dict(bold=True, color=NAVY)),
 ("  2 · Module - Feature - Permission — every product module, the features it includes, and the v2 permission tag behind each; plus which roles/grants hold it by default and which legacy DB tags map into it (blank = net-new in v2).", {}),
 ("  3 · Role x Permission — the 5 fixed roles as a tick matrix over all 51 v2 permissions.", {}),
 ("  4 · Grant x Permission — the 4 admin grants the same way. Global Admin is the superuser: every set, including the 3 exception permissions.", {}),
 ("  5 · Legacy tag audit — all 121 permission tags from the production database, each mapped to its v2 home, with every tag that does NOT carry into v2 flagged (deleted / retired / platform baseline).", {}),
 ("", {}),
 ("Reading notes:", dict(bold=True, color=NAVY)),
 ("  • The 3 exception permissions (remote.actuate, flags.impersonate, flags.sensorhealth) are in NO role. They are granted one person / one plant / one written reason at a time — Global Admin alone holds them implicitly.", {}),
 ("  • A ✓ in the matrices is the DEFAULT. Per-person overrides can add or remove any line at a plant, with a mandatory reason — that is the only flexibility mechanism.", {}),
 ("  • Every permission still needs its module licensed at the plant to actually work (see the Module column on each sheet).", {}),
 ("  • Floc Detector is a hardware entitlement with zero permissions — it appears on sheet 2 for completeness.", {}),
 ("", {}),
 ("Sources: index.html (SETS/ROLES/GRANTS/MODULES — extracted live), coverage-map.csv (the authoritative 121-tag mapping), data/Permissions.json + Features.json + Modules.json (production DB export, local-only — kept out of the public repo).", dict(size=9, color=GREY)),
 ("Deprecation-flag detail (which kept tags sit behind an L-* sunset flag) lives in reference/role-permission-migration-map.xlsx.", dict(size=9, color=GREY)),
]
for i, (txt, fkw) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = F(**fkw) if fkw else F()
    c.alignment = WRAP

# ---------------------------------------------- 2 · Module - Feature - Permission
ws = wb.create_sheet("Module - Feature - Permission")
headers = ["Product module", "Feature it includes", "v2 permission tag", "Permission (plain language)",
           "Permission set", "In roles by default", "In grants", "Legacy DB tags mapped here"]
ws.append(headers)
style_header(ws, 1, len(headers))
set_widths(ws, [24, 38, 24, 52, 20, 26, 30, 46])
ws.freeze_panes = "A2"

r = 2
for mod_id in ["core", "ops", "tasks", "data", "analytics", "iot", "inv", "floc"]:
    mod = mod_by_id[mod_id]
    mod_label = "%s %s (%s)%s" % (mod["icon"], mod["n"], mod_id, " — always included" if mod.get("always") else "")
    if mod.get("noperm"):
        ws.append([mod_label, "Hardware add-on — floc detection at the plant", "—",
                   "No user permissions. Plant-level entitlement only, so contracts track what hardware was sold; never caps or grants anything user-side.",
                   "—", "—", "—", "—"])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c); cell.font = F(italic=True, color=GREY); cell.alignment = WRAP; cell.border = THIN
        ws.cell(row=r, column=1).font = F(bold=True, color=NAVY)
        r += 1
        continue
    for feat, tags in FEATURES[mod_id]:
        for i, tag in enumerate(tags):
            roles = ", ".join(ROLES[x]["n"] for x in ROLE_ORDER if in_role(x, tag)) or "— none (exception grant: one person, one plant, one reason)"
            grants = ", ".join(GRANTNAMES[g] for g in GRANT_ORDER if in_grant(g, tag))
            legacy = " · ".join(legacy_sources.get(tag, [])) or "— net-new in v2 (no legacy tag)"
            ws.append([mod_label if (feat == FEATURES[mod_id][0][0] and i == 0) else "",
                       feat if i == 0 else "", tag, perm_name(tag), set_by_id[tag.split(".")[0]]["n"],
                       roles, grants, legacy])
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c); cell.font = F(); cell.alignment = WRAP; cell.border = THIN
            ws.cell(row=r, column=1).font = F(bold=True, color=NAVY)
            if i == 0:
                ws.cell(row=r, column=2).fill = FEAT_FILL
                ws.cell(row=r, column=2).font = F(bold=True)
            ws.cell(row=r, column=3).font = F(name="Courier New", size=9)
            if tag in ("remote.actuate", "flags.impersonate", "flags.sensorhealth"):
                for c in range(2, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = FLAG_FILL
            r += 1
total_row = r + 1
ws.cell(row=total_row, column=4, value="v2 permissions listed (formula):").font = F(bold=True)
ws.cell(row=total_row, column=3, value='=COUNTIF(C2:C%d,"*.*")' % (r - 1)).font = F(bold=True)
ws.cell(row=total_row + 1, column=4,
        value="Expected 51 — from index.html SETS (10 sets). 'Legacy DB tags mapped here' comes from coverage-map.csv; blank/net-new lines are the genuinely new capabilities (approvals, oversight co-sign, remote actuation, data correction, store setup, sensor health…).").font = F(size=9, color=GREY)
ws.cell(row=total_row + 1, column=4).alignment = WRAP

# ------------------------------------------------------------- 3/4 · matrices
def matrix_sheet(title, cols, member_fn, note):
    ws = wb.create_sheet(title)
    headers = ["Permission set", "v2 permission tag", "Permission (plain language)", "Product module"] + \
              [cols_n for cols_n in cols.values()]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    set_widths(ws, [22, 24, 50, 20] + [16] * len(cols))
    ws.freeze_panes = "E2"
    r = 2
    for s in SETS:
        for p in s["perms"]:
            tag = s["id"] + "." + p["id"]
            row = [s["n"], tag, p["n"], mod_by_id[PERMMOD[tag]]["n"]]
            row += ["✓" if member_fn(k, tag) else "—" for k in cols]
            ws.append(row)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c); cell.font = F(); cell.alignment = WRAP; cell.border = THIN
                if c > 4:
                    cell.alignment = CENTER
                    if cell.value == "—":
                        cell.font = F(color="C7CFD8")
            ws.cell(row=r, column=2).font = F(name="Courier New", size=9)
            if s["id"] in ("flags",) or s.get("sensitive"):
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = FLAG_FILL
            r += 1
    ws.cell(row=r + 1, column=3, value="Permissions granted by default (formula per column):").font = F(bold=True)
    for i in range(len(cols)):
        col = get_column_letter(5 + i)
        ws.cell(row=r + 1, column=5 + i, value='=COUNTIF(%s2:%s%d,"✓")' % (col, col, r - 1)).font = F(bold=True)
        ws.cell(row=r + 1, column=5 + i).alignment = CENTER
    note_cell = ws.cell(row=r + 3, column=1, value=note)
    note_cell.font = F(size=9, color=GREY); note_cell.alignment = WRAP
    ws.merge_cells(start_row=r + 3, start_column=1, end_row=r + 3, end_column=len(headers))
    return ws

matrix_sheet("Role x Permission",
             {k: ROLES[k]["n"] for k in ROLE_ORDER},
             in_role,
             "One role per person, account-wide (2026-07-30). A ✓ is the role's default; every line still needs its module licensed at the plant. "
             "The amber rows (IoT remote control, exception flags) are in NO role — they are per-person exception grants. "
             "Per-person overrides may add/remove lines at a plant with a written reason; dependency rules: approve needs work, oversight needs approve, portfolio needs readplant.")

matrix_sheet("Grant x Permission",
             {k: GRANTNAMES[k] for k in GRANT_ORDER},
             in_grant,
             "A grant sits ON TOP of the role and adds administration only; max one grant per person. Full Site = People + Technical. "
             "Global Admin is the superuser (2026-07-22): all 10 sets at every plant, INCLUDING the 3 exception permissions — the one holder for whom they are automatic. "
             "There are no custom roles (removed 2026-08-13): these four grants plus the five roles are the entire vocabulary.")

# ------------------------------------------------------------ 5 · Legacy tag audit
ws = wb.create_sheet("Legacy tag audit")
headers = ["Legacy permission tag (DB)", "DB name", "Legacy module (DB)", "Legacy feature (DB)",
           "v2 home", "Carried into v2?", "Left-out category", "Ruling / note (coverage-map.csv)"]
ws.append(headers)
style_header(ws, 1, len(headers))
set_widths(ws, [36, 30, 24, 24, 22, 15, 30, 70])
ws.freeze_panes = "A2"

db_by_tag = {p["permissionTag"]: p for p in (DB_PERMS or [])}

def category(home, note):
    if home in PERMMOD:
        return "YES", ""
    if home == "DELETED":
        return "NO — LEFT OUT", "Feature deleted outright (owner ruling — see note)"
    if home == "RETIRED":
        return "NO — LEFT OUT", "Retired / superseded — no v2 permission"
    if home == "BASELINE":
        return "NO — LEFT OUT", "Platform baseline for all users — no permission needed"
    return "NO — LEFT OUT", "Unrecognized home: " + home

r = 2
for row in COVER:
    tag, home, note = row["Production permission"], row["V2 home"], row["Status/Note"]
    db = db_by_tag.get(tag, {})
    carried, cat = category(home, note)
    ws.append([tag, db.get("name", "—"),
               DB_MODS.get(db.get("moduleTag", ""), row["Module"]),
               DB_FEATS.get(db.get("featureTag", ""), "—"),
               home, carried, cat, note])
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c); cell.font = F(); cell.alignment = WRAP; cell.border = THIN
    ws.cell(row=r, column=1).font = F(name="Courier New", size=9)
    ws.cell(row=r, column=5).font = F(name="Courier New", size=9)
    if carried != "YES":
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = FLAG_FILL
        ws.cell(row=r, column=6).font = F(bold=True, color=AMBER_INK)
        ws.cell(row=r, column=7).font = F(color=AMBER_INK)
    r += 1
last = r - 1

s = r + 1
summary = [
 ("Legacy DB tags audited (formula)", '=COUNTA(A2:A%d)' % last),
 ("Carried into a v2 permission (formula)", '=COUNTIF(F2:F%d,"YES")' % last),
 ("⚠ LEFT OUT of v2 — flagged amber above (formula)", '=COUNTIF(F2:F%d,"NO — LEFT OUT")' % last),
]
for label, formula in summary:
    ws.cell(row=s, column=1, value=label).font = F(bold=True)
    ws.cell(row=s, column=2, value=formula).font = F(bold=True)
    s += 1

# cross-checks computed at generation time (set differences are not expressible as sheet formulas)
db_tags = set(db_by_tag) if db_by_tag else None
csv_tags = {row["Production permission"] for row in COVER}
xdb = ("DB export not present at generation — audit built from coverage-map.csv alone (its 121 rows were pinned 1:1 against the DB on 2026-07-14)."
       if db_tags is None else
       "Cross-check at generation: DB tags missing from coverage-map.csv = %d; coverage-map rows not in the DB = %d (both must be 0)."
       % (len(db_tags - csv_tags), len(csv_tags - db_tags)))
c = ws.cell(row=s + 1, column=1, value=xdb + " Left-out tags stay visible in old data until their L-* deprecation flag is flipped (see role-permission-migration-map.xlsx); they never become v2 permissions.")
c.font = F(size=9, color=GREY); c.alignment = WRAP
ws.merge_cells(start_row=s + 1, start_column=1, end_row=s + 2, end_column=8)

wb.save(OUT)
db_note = DB_DIR or "not found (CSV-only mode)"
print(json.dumps({"written": OUT, "v2_perms": len(ALL_TAGS), "legacy_rows": len(COVER),
                  "carried": sum(1 for row in COVER if row["V2 home"] in PERMMOD),
                  "left_out": sum(1 for row in COVER if row["V2 home"] not in PERMMOD),
                  "db_export": db_note}))
